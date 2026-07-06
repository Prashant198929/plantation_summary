# Builds an Excel report of accounts that share a mobile number (almost
# always a leftover placeholder value from the sevakdb migration), so an
# admin can go collect each person's real number and fix their record.
#
# Usage: python build_dup_mobile_report.py <input_json_path> <output_xlsx_path>
import sys
import json
import openpyxl
from openpyxl.styles import Font, PatternFill

input_path = sys.argv[1]
output_path = sys.argv[2]

with open(input_path, "r", encoding="utf-8") as f:
    rows = json.load(f)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Duplicate_Mobile_Accounts"

headers = [
    "Group", "SharedValueMembers", "PlaceholderMobileValue",
    "Name_en", "Name_mr", "Zone", "BaithakPlace", "DocId", "Source",
]
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True)

band_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
last_group = None
band_on = False
for r in rows:
    if r["groupNum"] != last_group:
        band_on = not band_on
        last_group = r["groupNum"]
    ws.append([
        r["groupNum"],
        r["memberCount"],
        r["sharedMobilePlaceholder"],
        r["name"],
        r["name_mr"],
        r["zone"],
        r["baithakPlace"],
        r["docId"],
        r["source"],
    ])
    if band_on:
        for cell in ws[ws.max_row]:
            cell.fill = band_fill

for col, width in zip("ABCDEFGHI", [8, 16, 20, 26, 26, 12, 22, 20, 18]):
    ws.column_dimensions[col].width = width

wb.save(output_path)
print(f"Wrote {len(rows)} rows across {last_group} groups to {output_path}")
